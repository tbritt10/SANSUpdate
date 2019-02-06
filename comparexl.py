from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl

#Global strings for file names
sans = "Report1548267424539.xlsx"
active = "ActiveEEwdepart.xlsx"
report = "Employees.xlsx"
output = "report.xlsx"

audit = "CurriculumAuditReport.xlsx"
auditReport = "ActiveNotComplete.xlsx"

#Global variables for columns
sansEmail = "E"
sansFName = "C"
sansLName = "B"
sansEmpNum = "A"

activeEmail = "D"
activeFName = "A"
activeLName = "B"
activeEmpNum = "C"
activeDepartment = "E"
activeSupervisor = "F"


#The audit file is special in that it has merged cells, so the first cell in the merged cell range should return the data
auditCurriculum = "A"
auditEmpNum = "D"
auditName = "F"
auditEmail = "I"
auditStatus = "M"

class IncompleteRow:
    """Row class that represents rows from the incomplete training report"""

    def __init__(self, curriculum, empNum, name, email, status):
        """Create an incomplete row containing the data from given cells"""
        self.curriculum = curriculum
        self.empNum = empNum
        self.name = name
        self.email = email
        self.status = status

    def __eq__(self, other):
        """Overrides the default == (equals) behavior, returns a boolean value"""
        return self.email == other.email

    def __ne__(self, other):
        """Overrides the default != (not equals) behavior, returns a boolean value"""
        return self.email != other.email

    def returnCurriculum(self):
        """Return the curriculum from an object, returns a string"""
        return self.curriculum

    def returnEmpNum(self):
        """Returns the employee number from an object, returns a string"""
        return self.empNum

    def returnName(self):
        """Returns the name from an object, returns a sting"""
        return self.name

    def returnEmail(self):
        """Returns an email from an object, returns a sting"""
        return self.email

    def returnStatus(self):
        """Returns a status from an object, returns a sting"""
        return self.status


class ActiveRow:
    """Row class represents 4 Cells from each row in an excel sheet"""

    def __init__(self,email, fname, lname, empnum, department=None, supervisor=None):
        """Create a ActiveRow containing cells from the given row"""
        self.email = email
        self.fname = fname
        self.lname = lname
        self.empnum = str(empnum).zfill(9)
        if department is None:
            department = "General"
        self.department = department
        if supervisor is None:
            supervisor = "N/A"
        self.supervisor = supervisor

    def __eq__(self, other):
        """Overrides the defualt == (equals) behavior, returns a boolean value"""
        return self.email == other.email

    def __ne__(self, other):
        """Overrides the default != (not equal) behavior, returns a boolean value"""
        return self.email != other.email

    def returnEmail(self):
        """Pull the email from a row, returns a string"""
        return self.email

    def returnFName(self):
        """Pull the first name from a row, returns a string"""
        return self.fname

    def returnLName(self):
        """Pull the last name string from a row, returns a string"""
        return self.lname

    def returnEmpNum(self):
        """Pulls the employee number from a row, returns a string"""
        return self.empnum

    def returnDepartment(self):
        """Pulls the department string from a row, or returns the default value; returns a string"""
        return self.department

    def returnSupervisor(self):
        """Pulls the supervisor string from a row, or returns the default value; returns a string"""
        return self.supervisor

    def toString(self):
        return "{}".format(self.email)

    def compareRow(self, row):
        """Compare the email string between two rows, returns a Boolean value"""
        self.row = row
        email = self.email
        email2 = row.returnEmail
        if (email == email2):
            return True
        else:
            return False

##InactiveRow functions##
def loadAudit():
    #Load workbook and active worksheet
    wb = load_workbook(audit)
    ws = wb.active
    #Create list to store IncompleteRow objects
    auditRows = []
    print("Loading data from " + audit)
    for row in range(4, ws.max_row+1):
        formCurriculum = "{}{}".format(auditCurriculum, row)
        formEmpNum = "{}{}".format(auditEmpNum, row)
        formName = "{}{}".format(auditName, row)
        formEmail = "{}{}".format(auditEmail, row)
        formStatus = "{}{}".format(auditStatus, row)
        temprows = IncompleteRow(ws[formCurriculum].value, ws[formEmpNum].value, ws[formName].value, ws[formEmail].value, ws[formStatus].value)
        auditRows.append(temprows)
    print("Loaded data from " + audit)
    count = 0
    for i in range(len(auditRows)):
        if auditRows[i].returnStatus() == "Incomplete":
            count += 1
        else:
            continue
    print(str(count))
    return auditRows

def findIncomplete():
    #Compares incomplete list to inactive users, generates new list
    listInactive = findInactiveEmails()
    listIncomplete = loadAudit()
    resultList = []

    for email in listIncomplete:
        if email not in listInactive:
            resultList.append(email)

    return resultList

def exportAudit():
    auditList = findIncomplete()

    #Create new blank workbook
    wb = Workbook()
    ws = wb.active

    i = 1

    print("Attempting to write users to file")
    for x in range(len(auditList)):
        if auditList[x].returnStatus() == "Incomplete":
            ws.cell(row=i, column=1).value = auditList[x].returnCurriculum()
            ws.cell(row=i, column=3).value = auditList[x].returnEmpNum()
            ws.cell(row=i, column=5).value = auditList[x].returnName()
            ws.cell(row=i, column=7).value = auditList[x].returnEmail()
            ws.cell(row=i, column=9).value = auditList[x].returnStatus()
            i += 1
        else:
            continue
    print("Write successful\n")
    print("Attempting to save file as " + auditReport)
    try:
        wb.save(auditReport)
        print(auditReport + " saved\n")
    except PermissionError:
        print("ERROR: Permission denied; Please close " + auditReport + " to run the script")
        exit()



##SANS Functions##
def loadSANS():
    #Workbook and worksheet to pull data from
    wb = load_workbook(sans, data_only=True)
    ws = wb.active
    #List to store emails from wb
    sansRows = []
    print("Loading data from " + sans)
    for row in range(3,ws.max_row+1):
        formEmail = "{}{}".format(sansEmail, row)
        formFName = "{}{}".format(sansFName, row)
        formLName = "{}{}".format(sansLName, row)
        formEmpNum = "{}{}".format(sansEmpNum, row)
        temprows = ActiveRow(ws[formEmail].value, ws[formFName].value, ws[formLName].value, ws[formEmpNum].value)
        #Adding the cell reference.value to list
        sansRows.append(temprows)
    print("Loaded data from " + sans + "\n")
    return sansRows

##ActiveEE Functions##

def loadActive():
    #Loads emails
    #Workbook and worksheet to pull data from
    wb = load_workbook(active, data_only=True)
    ws = wb.active
    activeRows = []
    print("Loading data from " + active)
    for row in range(2,ws.max_row+1):
        formEmail = "{}{}".format(activeEmail, row)
        formFName = "{}{}".format(activeFName, row)
        formLName = "{}{}".format(activeLName, row)
        formEmpNum = "{}{}".format(activeEmpNum, row)
        formDepartment = "{}{}".format(activeDepartment, row)
        formSupervisor = "{}{}".format(activeSupervisor, row)
        temprows = ActiveRow(ws[formEmail].value, ws[formFName].value, ws[formLName].value, ws[formEmpNum].value, ws[formDepartment].value, ws[formSupervisor].value)
        #Adding the cell reference.value to list
        activeRows.append(temprows)
    print("Loaded emails from " + active + "\n")
    return(activeRows)

##Data Functions

def findInactiveEmails():
    #Pulls list data from other functions and compares the data to find
    #Inactive and new users, creating a new list
    print("Searching for inactive users\n")
    listSANS = loadSANS()
    listActive = loadActive()
    listInactive = []

    for email in listSANS:
        if email not in listActive:
            listInactive.append(email)

    listInactive = [value for value in listSANS if value not in listActive]
    print("Finished searching\n")
    print("Number of inactive users found: " + str(len(listInactive)) + "\n")
    return(listInactive)

def findNewEmails():
    print("Searching for new users\n")
    listSANS = loadSANS()
    listActive = loadActive()
    listNew = []

    for email in listActive:
        if email not in listSANS:
            listNew.append(email)

    print("Finished searching\n")
    print("Number of new users found: " + str(len(listNew)) + "\n")
    return(listNew)

def exportData():
    inactiveEmails = findInactiveEmails()
    newEmails = findNewEmails()

    i = 2
    print("Attempting to open output file")
    try:
        wb = load_workbook(report)
        ws = wb.active
        print("Output file loaded successfully\n")
    except FileNotFoundError:
        print("\nERROR: " + report + " not found")
        print("ERROR: Please download the file from the SANS Administrator portal")
        print("ERROR: The file must be placed in the same directory as the program")
        exit()
    #Put Inactive users into file
    print("Attempting to write inactive users to file")
    for x in range(len(inactiveEmails)):
        ws.cell(row=i, column = 4).value = inactiveEmails[x].returnEmail()
        ws.cell(row=i, column = 7).value = "NO"
        ws.cell(row=i, column = 8).value = "NO"
        ws.cell(row=i, column = 3).value = inactiveEmails[x].returnFName()
        ws.cell(row=i, column = 2).value = inactiveEmails[x].returnLName()
        ws.cell(row=i, column = 1).value = inactiveEmails[x].returnEmpNum()
        i+=1
    print("Write successful\n")

    print("Attempting to write new users to file")
    for x in range(len(newEmails)):
        ws.cell(row=i, column = 4).value = newEmails[x].returnEmail()
        ws.cell(row=i, column = 7).value = "YES"
        ws.cell(row=i, column = 8).value = "YES"
        ws.cell(row=i, column = 3).value = newEmails[x].returnFName()
        ws.cell(row=i, column = 2).value = newEmails[x].returnLName()
        ws.cell(row=i, column = 1).value = newEmails[x].returnEmpNum()
        ws.cell(row=i, column = 11).value = newEmails[x].returnDepartment()
        ws.cell(row=i, column = 5).value = newEmails[x].returnSupervisor()
        i+=1
    print("Write successful\n")
    print("Saving file...")
    try:
        wb.save(output)
        print(output + " saved\n")
    except PermissionError:
        print("ERROR: Permission denied; Please close " + output + " to run the script")

def main():
    exportData()
    exportAudit()
    print("Press enter to close this window")
    input()
main()
